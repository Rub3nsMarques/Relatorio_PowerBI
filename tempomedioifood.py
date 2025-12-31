import os
from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, Field
from openpyxl import load_workbook

from app.automation.comandas.relatorio_comandas_tempomedio_ifood import run_automation_and_download
from app.utils.responses import erro_resposta
from app.utils.auth import token_auth
from app.utils.logger_config import configurar_logger
from app.utils.lock import file_lock

logger = configurar_logger()
router = APIRouter(prefix="/relatorios", tags=["Tempo médio ifood"])

# ------------------ MODELO DE ENTRADA ------------------
class PeriodoRequest(BaseModel):
    data_inicio: str = Field(..., description="Data e hora inicial no formato DD/MM/AAAA HH:MM")
    data_fim: str = Field(..., description="Data e hora final no formato DD/MM/AAAA HH:MM")
    hora: int | None = Field(None, description="Hora que irá referenciar o turno do período (opcional)")

# ------------------ ENDPOINT ------------------
@router.post(
    "/relatorio-tempomedioifood",
    dependencies=[Depends(token_auth)],
    responses={
        200: {"description": "JSON com dados do relatório"},
        400: {"description": "Parâmetros inválidos"},
        401: {"description": "Não autorizado"},
        404: {"description": "Relatório vazio"},
        429: {"description": "Já existe uma automação em execução"},
        500: {"description": "Erro interno"},
    }
)
def gerar_relatorio_tempomedioifood(payload: PeriodoRequest):
    try:
        data_inicio = payload.data_inicio
        data_fim = payload.data_fim
        hora = payload.hora

        logger.info(f"Iniciando automação TEMPO MÉDIO IFOOD: {data_inicio} até {data_fim} (Ref hora: {hora})")

        with file_lock("relatorio_tempomedio_ifood", timeout=900):
            arquivo_path = run_automation_and_download(
                datahora_inicio=data_inicio, 
                datahora_fim=data_fim, 
                turno=str(hora) if hora is not None else None
            )

        # ---  Ler o XLSX e converter para JSON ---
        if not os.path.exists(arquivo_path):
            raise HTTPException(status_code=500, detail="Arquivo gerado não encontrado.")

        dados_json = []
        try:
            wb = load_workbook(arquivo_path, data_only=True)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                item = {str(k): v for k, v in zip(headers, row) if k is not None}
                if item: 
                    dados_json.append(item)
                    
            logger.info(f"✅ Dados convertidos para JSON: {len(dados_json)} registros")
            
            os.remove(arquivo_path) 

        except Exception as e:
            logger.error(f"Erro ao ler XLSX para JSON: {e}")
            raise HTTPException(status_code=500, detail="Erro ao processar arquivo gerado.")

        # Retorna o JSON direto. O FastAPI converte automaticamente.
        return {
            "status": "sucesso",
            "count": len(dados_json),
            "data": dados_json
        }

    except FileNotFoundError as e:
        logger.warning(f"Relatório vazio: {e}")
        raise HTTPException(
            status_code=404, 
            detail=erro_resposta("O relatório não gerou dados ou o botão de exportação não apareceu.")
        )
    
    except TimeoutError:
        logger.warning("Conflito: Automação já em execução.")
        raise HTTPException(
            status_code=429,
            detail=erro_resposta("Já existe uma automação em execução. Tente novamente.")
        )

    except ValueError as e:
        raise HTTPException(status_code=400, detail=erro_resposta(str(e)))

    except Exception as e:
        logger.error(f"Erro inesperado: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=erro_resposta("Erro interno durante a automação."))